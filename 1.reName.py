'''
0、删除指定excel文件最后一行有效数据
1、清洗数据
2、重命名文件
3、根据文件【sys-yishupingtai.xlsx】, 筛选数据并生成【sys-nopay.xlsx】和【sys-ispay.xlsx】文件
4、生成的两个文件 , 与ems的两个文件进行数据匹配
5、处理匹配付邮和免邮订单的最终对齐文件生成
6、执行【订单中心订单_中山纪念图书馆】的数据
7、执行【订单中心订单_佛山市图书馆】的数据
8、匹配中图错入格口的订单数据
'''
import os
import shutil
import argparse
import time
from typing import Union
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# 全局忽略openpyxl警告 , 自定义开启或关闭
import warnings
from openpyxl.styles.stylesheet import Stylesheet
warnings.filterwarnings("ignore", category=UserWarning, module=Stylesheet.__module__)

# ====================== 功能0：删除Excel最后一行有效数据 ======================
def has_real_data(ws, row_num):
    """检测指定行是否包含真实数据（保留原有注释）"""
    for cell in ws[row_num]:
        if cell.value not in [None, "", " "]:  # 排除空值、空字符串和空格
            return True
    return False

def delete_last_row_enhanced(excel_paths: list, sheet_name: str = None):
    """
    删除Excel文件最后一行有效数据（增强版）
    参数说明：
        :param excel_paths: Excel文件路径列表（支持多个文件）
        :param sheet_name: 指定统一的工作表名称（可选）
    """
    for path in excel_paths:
        print("\n" + "="*40)
        print(f"🟢 开始处理 Excel 文件: {path}")
        # print(f"🔧 工作表: '{sheet_name}'" if sheet_name else "🔧 工作表: 使用第一个工作表（默认）") 

        try:
            # ========== 加载阶段 ==========
            load_start = time.time()
            print("\n[1/3] ⏳ 正在加载工作簿...", end=" ", flush=True)
            wb = load_workbook(path)
            sheet_names = wb.sheetnames
            
            # 获取目标工作表
            if sheet_name:
                if sheet_name not in sheet_names:
                    raise KeyError(f"工作表 '{sheet_name}' 不存在")
                ws = wb[sheet_name]
            else:
                ws = wb.worksheets[0]
                sheet_name = ws.title
            
            print(f"✅ 完成（耗时 {time.time()-load_start:.1f}s）")
            print(f"   ▸ 工作表列表: {', '.join(sheet_names)}")
            print(f"   ▸ 当前操作表: '{sheet_name}'")

            # ========== 数据检测阶段 ==========
            print("\n[2/3] 🔍 正在深度检测数据...")
            original_max_row = ws.max_row
            
            # 空表检测（兼容格式残留情况）
            if original_max_row == 0 or not any(has_real_data(ws, row) for row in range(1, original_max_row+1)):
                print("⏹️ 终止：工作表无有效数据")
                return
            # 定位最后有效行
            last_data_row = next(
                row for row in reversed(range(1, original_max_row+1))
                if has_real_data(ws, row)
            )
            print(f"   ▸ 最后有效数据行: 第 {last_data_row} 行")

            # ========== 删除操作阶段 ==========
            print(f"\n🗑️ 正在删除第 {last_data_row} 行...", end=" ", flush=True)
            ws.delete_rows(last_data_row)
            print("✅ 完成")
            # ========== 保存阶段 ==========
            print("\n[3/3] 💾 正在保存文件...", end=" ", flush=True)
            save_start = time.time()
            wb.save(path)
            print(f"✅ 完成（耗时 {time.time()-save_start:.1f}s）")
            # ========== 最终报告 ==========
            print("\n" + "="*40)
            print(f"🎉 操作成功！总耗时: {time.time()-load_start:.1f}秒")
            print(f"➜ 原行数: {original_max_row} → 新行数: {ws.max_row}")

        except StopIteration:
            print("\n⚠️ 异常：检测到空行但无法定位有效数据")
        except KeyError as e:
            print(f"\n❌ 错误：{str(e)}")
            print("可用工作表列表：")
            for idx, name in enumerate(sheet_names, 1):
                print(f"  {idx}. {name}")
        except FileNotFoundError:
            print("\n❌ 错误：文件不存在，请检查路径是否正确")
        except PermissionError:
            print("\n❌ 错误：文件被占用，请关闭Excel后重试")
        except Exception as e:
            print(f"\n❌ 文件 {path} 发生未知错误：{str(e)}")
            continue

# ====================== 功能1：数据清洗与增强 ======================
def clean_data(input_path, output_path=None, delete_original=False):
    """
    :param delete_original: 是否删除原文件（需显式启用）
    对文件【易书平台订单表.xlsx】的数据进行清洗与增强处理：
    1. 标准化列名（'收件市'一列处 去除首尾空格）
    2. 在「区域类型」后添加「区域标识」列
    """
    # 路径处理
    input_path = Path(input_path)
    output_path = Path(output_path) if output_path else input_path.with_name(f"result_cleaned_{input_path.name}")
    try:
        df = pd.read_excel(input_path)
        df.columns = [col.strip() for col in df.columns]
        # 如果区域标识列已存在 → 先删除
        if '区域标识' in df.columns:
            df.drop(columns=['区域标识'], inplace=True)
            print("检测到已有区域标识列，已执行覆盖更新")
        # 插入区域标识列（无论之前是否存在）
        if '区域类型' in df.columns:
            insert_pos = df.columns.get_loc('区域类型') + 1
            region_map = {'市内':2, '省内':1}
            df.insert(
                insert_pos,
                '区域标识',
                df['区域类型'].map(region_map).fillna(0).astype(int)
            )
        else:
            raise ValueError("必要列「区域类型」缺失")
        df.to_excel(output_path, index=False)
        print(f"清洗完成 → {output_path}")
        # 删除原文件（危险操作！）
        if delete_original and input_path.exists():
            os.remove(str(input_path))  # 直接删除，不可恢复！
            print(f"原文件已永久删除：{input_path.name}")

    except Exception as e:
        print(f"处理失败：{input_path.name} | 错误类型：{type(e).__name__} | 详情：{str(e)}")
        if output_path and Path(output_path).exists():
            os.remove(output_path)  # 清理无效结果文件

def cleaning():
    """数据清洗"""
    clean_data(input_path="易书平台订单表.xlsx", delete_original=True)

# ====================== 功能2：批量重命名文件 ======================
def batch_rename(file_mapping, target_dir="."):
    """
    批量重命名文件，处理文件名冲突
    :param file_mapping: 字典类型，格式 {旧文件名: 新文件名}
    :param target_dir: 文件所在目录（默认为当前目录）
    """
    for old_name, new_name in file_mapping.items():
        old_path = Path(target_dir) / old_name
        new_path = Path(target_dir) / new_name
        if not old_path.exists():
            print(f"⚠️ 文件不存在: {old_name} - 已跳过")
            continue
        # 处理文件名冲突（自动添加序号）
        counter = 1
        temp_new_path = new_path
        while temp_new_path.exists():
            stem = new_path.stem
            suffix = new_path.suffix
            temp_new_path = new_path.with_name(f"{stem}_{counter}{suffix}")
            counter += 1
        try:
            shutil.move(str(old_path), str(temp_new_path))
            print(f"✅ 重命名成功: {old_name} → {temp_new_path.name}")
        except Exception as e:
            print(f"❌ 重命名失败: {old_name} → {new_name} | 错误信息: {str(e)}")

def renameExcel():
    """文件批量重命名"""
    rename_mapping = {
        "3404 7月.xlsx": "ems-nopay-3404.xlsx",
        "3513 7月.xlsx": "ems-ispay-3513.xlsx",
        "订单中心数据表-爱阅有约或科图承接.xlsx": "sys-aiyueyouyue.xlsx",
        "订单中心数据表-易书承接.xlsx": "sys-yishuchenjie.xlsx",
        "result_cleaned_易书平台订单表.xlsx": "sys-yishupingtai.xlsx"
    }
    batch_rename(rename_mapping)

# ====================== 功能3：Excel数据筛选 （支持动态多文件处理）======================
def excel_like_filter(
    input_path: Union[str, list],  # 支持字符串或列表输入
    filter_dict: dict = None,
    keep_columns: list = None,
    enhanced: bool = False,
    output_prefix: str = "_result"  # 输出文件名后缀
):
    """    
    参数：
    input_path   - 输入路径（支持单个文件路径或文件路径列表）
    output_prefix - 输出文件名后缀 (默认 _result)

    示例：
    excel_like_filter("a.xlsx") → 生成 result_a.xlsx
    excel_like_filter(["a.xlsx", "b.xlsx"]) → 生成 result_a.xlsx 和 result_b.xlsx

    :param filter_dict: 筛选条件字典（格式见示例）
    :param keep_columns: 保留列列表
    :param enhanced: 是否启用增强模式（支持 != 多值排除等特性）

    基础模式条件示例：
    {'年龄': ('>', 30), '城市': ['北京','上海']}

    增强模式新增特性：
    - 支持 != 运算符排除多个值
    - 支持 not in 运算符
    - 更严格的类型检查
    """
 
 # 统一转换为文件路径列表
    if isinstance(input_path, str):
        input_files = [Path(input_path)]
    else:
        input_files = [Path(p) for p in input_path]

    # 处理每个文件
    for in_file in input_files:
        # 生成输出路径
        out_file = in_file.with_name(f"{in_file.stem}{output_prefix}.xlsx")
        
        try:
            # 读取Excel数据
            df = pd.read_excel(in_file)
            if filter_dict:
                mask = pd.Series(True, index=df.index)
                for col, condition in filter_dict.items():
                    # 执行数据筛选
                    if filter_dict:
                        mask = pd.Series(True, index=df.index)
                        for col, condition in filter_dict.items():
                            if col not in df.columns:
                                raise ValueError(f"列 '{col}' 不存在")
                                
                            # 处理元组条件（运算符+值）
                            if isinstance(condition, tuple):
                                operator, value = condition
                                # 增强模式：严格校验运算符
                                if enhanced and operator not in ['>','<','>=','<=','!=','in','not in']:
                                    raise ValueError(f"无效运算符：{operator}（增强模式要求使用明确运算符）")
                                # 处理多值排除（增强模式特性）
                                if operator == '!=' and isinstance(value, list):
                                    mask &= ~df[col].isin(value)
                                elif operator == '>':
                                    mask &= (df[col] > value)
                                elif operator == '<':
                                    mask &= (df[col] < value)
                                elif operator == '>=':
                                    mask &= (df[col] >= value)
                                elif operator == '<=':
                                    mask &= (df[col] <= value)
                                elif operator == '!=':
                                    mask &= (df[col] != value)
                                elif operator == 'in':
                                    mask &= df[col].isin(value)
                                elif operator == 'not in':
                                    mask &= ~df[col].isin(value)
                                else:
                                    if enhanced:
                                        raise ValueError(f"增强模式不支持运算符：{operator}")
                                    # 基础模式兼容旧写法
                                    mask &= (df[col] == condition)
                            # 处理列表条件（自动转换为IN操作）
                            elif isinstance(condition, list):
                                mask &= df[col].isin(condition)
                            # 处理单值条件
                            else:
                                mask &= (df[col] == condition)
                        df = df[mask]

            if keep_columns:
                # 处理列选择
                if keep_columns:
                    missing_cols = [col for col in keep_columns if col not in df.columns]
                    if missing_cols:
                        raise ValueError(f"缺失列：{missing_cols}")
                    df = df[keep_columns]

            # 保存结果
            df.to_excel(out_file, index=False)
            print(f"处理成功：{in_file.name} → {out_file.name}")
            
        except Exception as e:
            print(f"处理失败：{in_file.name} | 错误：{str(e)}")

# ===excel_like_filter函数对应的特定场景快捷调用 ===
def filter_pay_postage():
    """付邮筛选"""
    excel_like_filter(
        input_path="sys-yishupingtai.xlsx",
        output_prefix="_sys-ispay-1",
        filter_dict={
            '是否需要邮费': ['付邮'],
            '支付类型': ['微信']
        },
        keep_columns=['快递单号', '借还书订单号', '图书馆名称', '流水订单号', '订单创建时间','订单状态', '支付类型', '收件人详细地址', '收件人名称', '实付金额-单位为分','是否需要邮费', '收件市', '区域类型', '区域标识'],
        enhanced=False
    )

def filter_wechat_pay():
    """微信支付筛选 , 仅适用于：当场付邮筛选无数据时启用"""
    excel_like_filter(
        input_path="sys-yishuchenjie.xlsx",
        output_prefix="_sys-ispay-2",
        filter_dict={
            '支付类型': ['微信']
        },
        keep_columns=["借书者id", "快递单号", "承接应用方订单id", "承接应用名称",
        "所属图书馆名称", "支付订单号", "创建时间", "订单状态",
        "支付类型", "收件人地址", "收件人姓名", "应付金额"],
        enhanced=False
    )

def filter_free_postage():
    """免邮筛选"""
    excel_like_filter(
        input_path="sys-yishupingtai.xlsx",
        output_prefix="_sys-nopay",
        filter_dict={
            '是否需要邮费': ['免邮']
        },
        keep_columns=['快递单号', '借还书订单号', '图书馆名称', '订单创建时间',
        '是否需要邮费', '收件人名称', '收件市', '区域类型', '区域标识', '支付类型'],
        enhanced=False
    )

def filter_zhongshan_orders():
    """订单中心订单_中山纪念图书馆的数据筛选"""
    excel_like_filter(
        input_path=["sys-yishuchenjie.xlsx","sys-aiyueyouyue.xlsx"],
        output_prefix="_sys-zslib",
        filter_dict={
            '订单状态': ('!=', '已关闭'),
            '所属图书馆名称': ('!=', '佛山市图书馆')
        },
        keep_columns=['借书者id', '发起应用方订单id', '承接应用方订单id', '发起应用名称', '所属图书馆名称', '订单类型',  '订单状态', '借书人名称', '收件人地址', '创建时间', '应付金额', '支付类型'],
        enhanced=True  # 启用增强模式
    )

def filter_foshan_orders():
    """订单中心订单_佛山市图书馆的数据筛选"""
    excel_like_filter(
        input_path="sys-aiyueyouyue.xlsx",
        output_prefix="_sys-fslib",
        filter_dict={
            '订单状态': ('!=', '已关闭'),
            '订单类型': ['快递借书'], #原为【快递类型】
            '快递单号': ('!=', 0), #0为测试订单 , 不统计
            '所属图书馆名称': '佛山市图书馆'
        },
        keep_columns=['借书者id', '发起应用方订单id', '承接应用方订单id', '发起应用名称', '所属图书馆名称', '订单类型', '快递单号', '订单状态', '借书人名称', '收件人地址', '创建时间', '应付金额', '支付类型'],
        enhanced=True
    )

# ====================== 功能4：数据比对与标记 ======================
def compare_and_highlight_general(
    sys_file: str, 
    ems_file: str,
    output_sys_marked: str,
    output_ems_marked: str,
    columns_to_keep: list,
    invalid_replace: str = '无效单号'
):
    """
    通用数据比对与标记函数
    :param sys_file: 系统订单文件路径（如 sys-ispay.xlsx）
    :param ems_file: 快递数据文件路径（如 ems-ispay-3513.xlsx）
    :param output_sys_marked: 标黄后系统文件保存路径
    :param output_ems_marked: 标红后快递文件保存路径
    :param columns_to_keep: 需要保留的列列表
    :param invalid_replace: 无效单号替换文本
    """
    try:
        # 读取数据（强制使用openpyxl引擎）
        df_a = pd.read_excel(sys_file, dtype={'快递单号': str}, engine='openpyxl')
        df_b = pd.read_excel(ems_file, dtype={'快递单号': str}, engine='openpyxl')
    except Exception as e:
        print(f"文件读取失败: {str(e)}")
        return

    # 数据清洗
    df_a['快递单号'] = df_a['快递单号'].str.strip().replace(['0', ''], invalid_replace)
    df_b['快递单号'] = df_b['快递单号'].str.strip()

    # 获取关键数据集
    original_target_numbers = set(df_b['快递单号'])
    target_numbers = set(df_a['快递单号'])

    # 标黄操作
    wb_a = load_workbook(sys_file)
    ws_a = wb_a.active
    yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")
    matching_results = []
    
    # 标红操作
    wb_b = load_workbook(ems_file)
    ws_b = wb_b.active
    red_fill = PatternFill(start_color="FF0000", fill_type="solid")
    unmatched_results = []

    # 处理标黄
    for idx, row in df_a.iterrows():
        if row['快递单号'] in original_target_numbers:
            ws_a.cell(row=idx+2, column=1).fill = yellow_fill
            matching_results.append(row.to_dict())
            print(f"标黄：{row['快递单号']}")

    # 处理标红
    for idx, row in df_b.iterrows():
        if row['快递单号'] not in target_numbers:
            ws_b.cell(row=idx+2, column=1).fill = red_fill
            unmatched_results.append(row.to_dict())
            print(f"未匹配到：{row['快递单号']}")

    # 保存标记文件
    wb_a.save(output_sys_marked)
    wb_b.save(output_ems_marked)

    # 生成结果文件
    matching_df = pd.DataFrame(matching_results)
    if matching_df.empty:
        print("警告：没有匹配到任何记录！")
        return

    # 列存在性检查
    missing_columns = [col for col in columns_to_keep if col not in matching_df.columns]
    if missing_columns:
        print(f"错误：以下列不存在: {missing_columns}")
        print("当前数据列:", list(matching_df.columns))
        return

    # 生成最终结果
    try:
        matching_df = matching_df[columns_to_keep]
        matching_df.insert(0, "序号", range(1, len(matching_df)+1))
        
        # 动态生成结果文件名
        base_name = os.path.splitext(output_sys_marked)[0]
        matching_df.to_excel(f'{base_name}-匹配结果.xlsx', index=False)
        pd.DataFrame(unmatched_results).to_excel(f'{base_name}-未匹配结果.xlsx', index=False)
        
        print(f"\n处理结果：{os.path.basename(sys_file)}")
        print(f"✅ 标黄文件已保存: {output_sys_marked}")
        print(f"✅ 标红文件已保存: {output_ems_marked}")
        print(f"📊 匹配记录: {len(matching_df)} 条")
        print(f"📊 未匹配记录: {len(unmatched_results)} 条")
    except Exception as e:
        print(f"保存结果失败: {str(e)}")

# ===compare_and_highlight_general函数对应的特定场景快捷调用 ===
def compare_ispay():
    """处理付邮订单比对"""
    compare_and_highlight_general(
        sys_file='sys-ispay.xlsx',
        ems_file='ems-ispay-3513.xlsx',
        output_sys_marked='sys-ispay-marked.xlsx',
        output_ems_marked='ems-ispay-3513-marked.xlsx',
        columns_to_keep=[
            "快递单号", "承接应用方订单id", "所属图书馆名称",
            "支付订单号", "创建时间", "收件人地址", "收件人姓名",
            "借书者id", "应付金额"
        ]
    )

def compare_nopay():
    """处理免邮订单比对"""
    compare_and_highlight_general(
        sys_file='sys-nopay.xlsx',
        ems_file='ems-nopay-3404.xlsx',
        output_sys_marked='sys-nopay-marked.xlsx',
        output_ems_marked='ems-nopay-3404-marked.xlsx',
        columns_to_keep=[
            "快递单号", "借还书订单号", "图书馆名称",
            "订单创建时间", "是否需要邮费", "收件人名称",
            "收件市", "区域类型", "区域标识"
        ]
    )

def compare_error():
    """处理入中图错格口订单比对"""
    compare_and_highlight_general(
        sys_file='sys-yishupingtai.xlsx', # 统计类型为还书的订单 , 再进行匹配
        ems_file='ems-errordata.xlsx',
        output_sys_marked='sys-yishupingtai-marked.xlsx',
        output_ems_marked='ems-errordata-marked.xlsx',
        columns_to_keep=[
            "快递单号", "借还书订单号", "图书馆名称", "订单创建时间", "订单类型", "是否需要邮费",  "收件人名称", "收件市", "区域类型", "区域标识", "实付金额-单位为分", "流水订单号"
        ]
    )

# ====================== 功能5：生成最终付邮/免邮的对齐文件 ======================
# --- 全局配置 ---
"""
执行S8后，付邮订单数据最终匹配的数据 (即3513_佳禾月结对齐_xxx元_xxx月.xlsx)
"""
CONFIG_S8 = {
    'required_columns_a': ['快递单号', '所属图书馆名称', '支付订单号', '承接应用方订单id', '创建时间'],
    'a_columns_to_merge': ['快递单号', '所属图书馆名称', '支付订单号', '承接应用方订单id', '创建时间'],
    'b_rename_columns': {'寄件人': '图书馆'},
    'b_drop_columns': ['产品'],
    'b_new_columns': ['支付订单号', '创建时间', '订单号'],
    'new_column_order': [
        '序号', '订单号', '图书馆', '创建时间', '支付订单号',
        '快递单号', '寄达市名称', '大宗客户名称', '收寄时间', '计费重量(克)', '总邮资'
    ],
    'column_mapping': {
        '图书馆': '所属图书馆名称',
        '支付订单号': '支付订单号',
        '订单号': '承接应用方订单id',
        '创建时间': '创建时间'
    }
}

"""
执行S9后，免邮订单数据最终匹配的数据 (即3404_省图免邮对齐_xxx元_xxx月.xlsx)
"""
CONFIG_S9 = {
    'required_columns_a': ['快递单号', '图书馆名称', '借还书订单号', '订单创建时间'],
    'a_columns_to_merge': ['快递单号', '图书馆名称', '借还书订单号', '订单创建时间'],
    'b_rename_columns': {'寄件人': '图书馆'},
    'b_drop_columns': ['产品'],
    'b_new_columns': ['创建时间', '订单号'],
    'new_column_order': [
        '序号', '订单号', '图书馆', '创建时间', '快递单号',
        '寄达市名称', '大宗客户名称', '收寄时间', '计费重量(克)', '总邮资'
    ],
    'column_mapping': {
        '图书馆': '图书馆名称',
        '订单号': '借还书订单号',
        '创建时间': '订单创建时间'
    }
}

# --- 处理对齐函数 ---
def process_excel_files(
    file_a_path: str,
    file_b_path: str,
    config: dict,
    output_path: str = "修改后的B文件.xlsx"
) -> dict:
    """
    处理Excel文件的主函数
    
    参数：
    file_a_path -- A文件路径（如sys-*-marked-匹配结果.xlsx）
    file_b_path -- B文件路径（如ems-*-marked.xlsx）
    config -- 处理配置字典（使用CONFIG_S8/CONFIG_S9）
    output_path -- 输出文件路径（默认：修改后的B文件.xlsx）
    
    返回：
    {
        "status": "success"|"error",
        "message": 描述信息,
        "output_path": 输出文件路径
    }
    """
    try:
        # ====== 读取文件 ======
        df_a = pd.read_excel(file_a_path)
        df_b = pd.read_excel(file_b_path)

        # ====== 校验必要列 ======
        required_columns = {
            'A文件': config['required_columns_a'],
            'B文件': ['序号', '产品', '快递单号', '寄件人', '寄达市名称', '大宗客户名称', '收寄时间', '计费重量(克)', '总邮资']
        }
        
        # 检查A文件列
        missing_a = [col for col in required_columns['A文件'] if col not in df_a.columns]
        if missing_a:
            raise ValueError(f"A文件缺少必要列：{missing_a}")

        # 检查B文件列
        missing_b = [col for col in required_columns['B文件'] if col not in df_b.columns]
        if missing_b:
            raise ValueError(f"B文件缺少必要列：{missing_b}")

        # ====== 处理B文件结构 ======
        # 1. 重命名列
        df_b = df_b.rename(columns=config['b_rename_columns'])
        
        # 2. 删除指定列
        df_b = df_b.drop(columns=config['b_drop_columns'])
        
        # 3. 插入新列
        for col in config['b_new_columns']:
            df_b[col] = ''

        # 4. 调整列顺序
        df_b = df_b[config['new_column_order']]

        # ====== 合并数据 ======
        merged = pd.merge(
            df_b[['快递单号']],  # 保留B文件的快递单号顺序
            df_a[config['a_columns_to_merge']],
            on='快递单号',
            how='left'
        )

        # ====== 数据映射 ======
        for b_col, a_col in config['column_mapping'].items():
            df_b[b_col] = merged[a_col]

        # ====== 保存结果 ======
        df_b.to_excel(output_path, index=False)
        
        return {
            "status": "success",
            "message": f"文件处理完成，已保存到：{output_path}",
            "output_path": output_path
        }

    except FileNotFoundError as e:
        return {
            "status": "error",
            "message": f"文件不存在：{str(e)}",
            "output_path": None
        }
    except ValueError as e:
        return {
            "status": "error",
            "message": str(e),
            "output_path": None
        }
    except KeyError as e:
        return {
            "status": "error",
            "message": f"列名错误：{str(e)}",
            "output_path": None
        }
    except Exception as e:
        return {
            "status": "error",
            "message": f"未知错误：{str(e)}",
            "output_path": None
        }

# --- 任务执行器 ---
def run_processing_task(task_config: dict) -> dict:
    """
    执行单个处理任务的统一入口
    
    参数：
    task_config -- 包含以下键的字典：
        name: 任务名称
        config: 处理配置（CONFIG_S8/CONFIG_S9）
        file_a: A文件路径
        file_b: B文件路径
        output: 输出文件路径
    
    返回：
    包含任务执行结果的字典
    """
    start_time = time.time()
    print(f"\n🔨 开始处理【{task_config['name']}】...")
    
    result = process_excel_files(
        file_a_path=task_config['file_a'],
        file_b_path=task_config['file_b'],
        config=task_config['config'],
        output_path=task_config['output']
    )
    
    # 统一结果输出格式
    status_icon = "✅" if result['status'] == "success" else "❌"
    elapsed_time = time.time() - start_time
    
    print(f"{status_icon} 任务状态：{result['status'].upper()}")
    print(f"📝 提示信息：{result['message']}")
    print(f"⏱ 处理耗时：{elapsed_time:.2f}秒")
    
    if result['status'] == "success":
        print(f"📂 输出位置：{result['output_path']}")
    
    return result

# ====================== 执行控制核心 ======================
class PipelineController:
    def __init__(self, cmd_args):
        self.args = cmd_args  # 保存命令行参数
        self.step_functions = {
            'S0': self.dynamic_delete_step, # 删除指定文件最后一行有效数据
            'S1': cleaning, # 数据清洗
            'S2': renameExcel, # 文件重命名
            'S3': filter_pay_postage, # 付邮筛选
            'S4': filter_wechat_pay, # 微信支付筛选 , 仅适用于：当S3筛选无数据时启用
            'S5': filter_free_postage, # 免邮筛选
            'S6': filter_zhongshan_orders, # 中山纪念图书馆的数据筛选
            'S7': filter_foshan_orders, # 佛山市图书馆的数据筛选
            'S8': compare_ispay, # 匹配付邮订单
            'S9': compare_nopay, # 匹配免邮订单
            'S10': compare_error, # 匹配错入格口订单
            'S11': self.process_s8_final,  # 匹配付邮最终处理
            'S12': self.process_s9_final   # 匹配免邮最终处理
        }

    def process_s8_final(self):
            """处理付邮订单的最终对齐文件"""
            task_config = {
                "name": "S8付邮订单",
                "config": CONFIG_S8,
                "file_a": "sys-ispay-marked-匹配结果.xlsx",
                "file_b": "ems-ispay-3513-marked.xlsx",
                "output": "ems-ispay-3513-marked-end-result.xlsx"
            }
            run_processing_task(task_config)

    def process_s9_final(self):
        """处理免邮订单的最终对齐文件"""
        task_config = {
            "name": "S9免邮订单",
            "config": CONFIG_S9,
            "file_a": "sys-nopay-marked-匹配结果.xlsx",
            "file_b": "ems-nopay-3404-marked.xlsx",
            "output": "ems-nopay-3404-marked-end-result.xlsx"
        }
        run_processing_task(task_config)

    def dynamic_delete_step(self):
            """动态删除步骤（实例方法）"""
            # 优先级处理：命令行参数 > 预设配置
            files = self.args.del_files or ["3513 7月.xlsx", "3404 7月.xlsx"]
            sheet = self.args.del_sheet  # 允许为None
            
            print(f"\n🔧 删除行配置：")
            print(f"   ▸ 目标文件: {files}")
            print(f"   ▸ 指定工作表: {sheet if sheet else '自动选择首表'}")

            delete_last_row_enhanced(
                excel_paths=files,
                sheet_name=sheet
            )

    def execute_pipeline(self, steps):
        """执行流水线操作"""
        print(f"🏁 开始执行流程：{' → '.join(steps)}")
        for step in steps:
            if step in self.step_functions:
                print(f"\n🔧 正在执行 {step}")
                self.step_functions[step]()  # 调用绑定方法
            else:
                print(f"⚠️ 未知步骤：{step}")
        print("\n✅ 所有指定步骤执行完成")

# ====================== 主程序执行示例 ======================
if __name__ == "__main__":

    # 配置命令行参数
        parser = argparse.ArgumentParser(description="数据处理流水线控制器（支持删除行操作）")
        parser.add_argument('-s', '--steps', 
                            nargs='+',
                            # default=['S0',S1','S2','S3','S4','S5','S8','S9','S6','S7','S10','S11','S12'], # 默认按此顺序自动执行
                            required=True,  # 必填参数
                            help="指定执行步骤序列（默认顺序：S0 S1 S2 S3 S4 S5 S8 S9 S6 S7 S10 S11 S12）")
        parser.add_argument('--del-files', nargs='+', 
                       help="指定需要删除行的文件列表")
        parser.add_argument('--del-sheet', 
                        help="指定统一工作表名称（可选）")
        args = parser.parse_args()
        # 创建控制器并执行
        controller = PipelineController(args)  # 注入命令行参数
        controller.execute_pipeline(args.steps)

'''
删除最后一行有效数据执行方法（详细）
======
py reName.py -s S0 # 预设配置 默认第一个工作表
py reName.py -s S0 --del-sheet "listpro" # 预设配置 指定工作表
py reName.py -s S0 --del-files "3513 4月.xlsx" --del-sheet "citylist" #指定文件指定工作表
py reName.py -s S0 --del-files "3513 4月.xlsx" "3404 4月.xlsx" --del-sheet "工作表名" #自定义文件自定义工作表 (--del-sheet不填默认首表)
======

# 邮政提供的订单表数据 删除最后一行
py reName.py -s S0
py reName.py -s S0 --del-files "3513 4月.xlsx" "3404 4月.xlsx" --del-sheet "工作表名"

# 易书平台订单表数据清洗/原始订单表数据批量重命名
py reName.py -s S1 S2

# 付邮/免邮筛选 生成sys-ispay.xlsx/sys-nopay.xlsx文件
py reName.py -s S3 S4 S5

# 匹配付邮订单数据
py reName.py -s S8
ems-ispay-3513-marked.xlsx
sys-ispay-marked-匹配结果.xlsx

# 匹配免邮订单数据
py reName.py -s S9
ems-nopay-3404-marked.xlsx
sys-nopay-marked-匹配结果.xlsx

# 生成最终的付邮/免邮订单数据
py reName.py -s S11
py reName.py -s S12

# 中山纪念馆/佛山图书馆订单数据筛选【给省图的免邮订单 细节处理待调整】
py reName.py -s S6 S7
sys-yishuchenjie_sys-zslib.xlsx
sys-aiyueyouyue_sys-zslib.xlsx
sys-aiyueyouyue_sys-fslib.xlsx

# 匹配错入格口订单数据
py reName.py -s S10
ems-errordata-marked.xlsx
sys-yishupingtai-marked-匹配结果.xlsx

'''





