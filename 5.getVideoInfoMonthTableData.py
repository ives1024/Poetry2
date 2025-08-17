import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
import calendar
import datetime

def create_excel_with_sheets():
    # 创建新工作簿
    wb = Workbook()
    year = 2025
    
    # 定义列名
    columns = ['ID', '抖音', '西瓜', '微信视频号', '爱奇艺', '优酷号', '支付宝生活号', '快手', '美团视频', 'B站', '皮皮虾', '小红书', '百家号', '好看视频', '腾讯视频', '腾讯微视', '多多视频']
    
    # 定义样式配置
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=False
    )
    header_font = Font(bold=True)
    weekend_fill = PatternFill(
        start_color='ADD8E6',
        end_color='ADD8E6',
        fill_type='solid'
    )

    # 创建各月份工作表
    for month in range(1, 13):
        if month == 1:
            ws = wb.active
            ws.title = f"{year}年{month}月"
        else:
            ws = wb.create_sheet(f"{year}年{month}月")
        
        # ========== 设置标题行样式 ==========
        # 添加标题行
        ws.append(columns)
        
        # 遍历标题行每个单元格
        for col in range(1, len(columns)+1):
            cell = ws.cell(row=1, column=col)
            
            # 设置字体样式
            cell.font = header_font
            
            # 设置对齐方式
            cell.alignment = center_alignment
            
            # 添加边框
            cell.border = thin_border
        
        # 设置标题行高
        ws.row_dimensions[1].height = 22
        
        # ========== 生成数据行 ==========
        _, days_in_month = calendar.monthrange(year, month)
        
        for day in range(1, days_in_month + 1):
            row_num = day + 1  # 数据从第2行开始
            
            # 添加数据行
            ws.append([day] + [''] * (len(columns)-1))
            
            # 设置行高
            ws.row_dimensions[row_num].height = 22
            
            # 处理每个单元格
            current_date = datetime.date(year, month, day)
            for col in range(1, len(columns)+1):
                cell = ws.cell(row=row_num, column=col)
                
                # 设置通用样式
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # 设置周末样式（仅ID列）
                if col == 1 and current_date.weekday() >= 5:
                    cell.fill = weekend_fill
        # ========== 冻结首行 ==========
        ws.freeze_panes = 'A2'

    # 保存文件
    wb.save(f"{year}自媒体数据报告_最终版.xlsx")
    print(f"文件已生成，所有样式配置正确应用！")

if __name__ == "__main__":
    create_excel_with_sheets()


